"""
Iterate docs/call_combinations.xlsx in batches of 10 rows:
  1. Update each of the 10 dummy agents' agent_config with the row's
     llm/stt/tts service_provider_id + model_id (written into metadata).
  2. Prompt the user to manually place the 10 calls in parallel.
  3. Poll call_logs until 10 completed calls (one per agent) are found.
  4. Write call_log_id + status='completed' back into the excel rows and save.
  5. Repeat until the sheet is exhausted.
"""
import json
import os
import time
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras
import redis
import requests

BASE_URL = "http://localhost:8000/api/v1"
AUTH_TOKEN = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoyLCJlbWFpbCI6InRoaWxhay5ndW5hc2VrYXJhbkBwcm9kdWN0ZnVzaW9uLmNvIiwib3JnX2lkIjoiOTRkMjgxMzgtYjU0MC00Mzg5LThkYzgtOWIzYmFlZjMyYmQ5Iiwicm9sZSI6Im93bmVyIiwiaWF0IjoxNzc1NzM2OTg0LCJleHAiOjE3NzU4MjMzODR9.CceVHS1OVpChC8D8cspB22dNM_-RQCIn4ZZJQR2snfU"
DB_URL = "postgresql://neondb_owner:npg_XFStP7eh4jZY@ep-blue-truth-an3xdo3i-pooler.c-6.us-east-1.aws.neon.tech/neondb?sslmode=require"
REDIS_URL = "redis://localhost:6379/0"

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "docs" / "call_combinations.xlsx"
AGENTS_JSON = Path(__file__).parent / "dummy_agent_ids.json"

BATCH_SIZE = 10
POLL_INTERVAL_S = 60
POLL_TIMEOUT_S = 10 * 60

HEADERS = {
    "Authorization": f"Bearer {AUTH_TOKEN}",
    "Content-Type": "application/json",
}

# Excel column indexes (1-based)
COL = {
    "llm_sp_id": 1,
    "llm_model_id": 3,
    "stt_sp_id": 5,
    "stt_model_id": 7,
    "tts_sp_id": 9,
    "tts_model_id": 11,
    "call_log_id": 13,
    "status": 14,
    "metrics": 15,
    "transcription": 16,
    "to_number": 17,
}


def load_agents() -> list[dict]:
    with open(AGENTS_JSON) as f:
        data = json.load(f)
    if len(data) < BATCH_SIZE:
        raise RuntimeError(f"Need {BATCH_SIZE} agents, found {len(data)} in {AGENTS_JSON}")
    return data[:BATCH_SIZE]


def fetch_existing_config(agent_id: int) -> dict:
    r = requests.get(
        f"{BASE_URL}/agent/get_agent",
        params={"agent_id": agent_id},
        headers=HEADERS,
    )
    r.raise_for_status()
    data = r.json()
    if isinstance(data, list):
        data = data[0] if data else {}
    # The agent response includes config fields at top level
    return data


def flush_redis_cache(agent_ids: list[int]) -> None:
    """Delete all agent_bot_data cache keys for the given agents."""
    try:
        r = redis.from_url(REDIS_URL)
        deleted = 0
        for agent_id in agent_ids:
            keys = r.keys(f"agent_bot_data:{agent_id}:*")
            if keys:
                deleted += r.delete(*keys)
        if deleted:
            print(f"  Flushed {deleted} Redis cache keys")
    except Exception as e:
        print(f"  WARNING: Redis flush failed: {e}")


def get_default_voice_for_tts(tts_sp_id: int) -> str | None:
    """Query the voices table to find a default voice_id for a given TTS provider."""
    try:
        conn = psycopg2.connect(DB_URL)
        with conn.cursor() as cur:
            cur.execute(
                "SELECT voice_id FROM voices WHERE service_provider_id = %s LIMIT 1",
                (tts_sp_id,),
            )
            row = cur.fetchone()
        conn.close()
        return row[0] if row else None
    except Exception:
        return None


def update_agent_config(agent_id: int, row: dict, base_config: dict) -> None:
    # Build tts_metadata with voice_id from voices table
    tts_meta = {"model_id": row["tts_model_id"]}
    voice_id = get_default_voice_for_tts(row["tts_sp_id"])
    if voice_id:
        tts_meta["voice_id"] = voice_id

    payload = {
        "agent_id": agent_id,
        "llm_service_id": row["llm_sp_id"],
        "stt_service_id": row["stt_sp_id"],
        "tts_service_id": row["tts_sp_id"],
        "llm_metadata": {"model_id": row["llm_model_id"]},
        "stt_metadata": {"model_id": row["stt_model_id"]},
        "tts_metadata": tts_meta,
        "system_prompt": "You are a polite and professional hotel receptionist. When the caller calls, greet them warmly and introduce the hotel. Help the caller with room booking enquiries. Ask the following questions one by one, in a natural conversation: 1. Check-in date 2. Check-out date 3. Number of guests 4. Type of room needed (single, double, deluxe, etc.). If the caller asks about price, give a reasonable estimate and clearly say that final pricing will be confirmed by the hotel staff. If the caller wants to make a booking, collect their full name, phone number, and email address",
        "first_message": base_config.get("first_message") or "Hello",
        "end_call_message": base_config.get("end_call_message"),
        "status": "active",
        "agent_metadata": base_config.get("agent_metadata") or {},
    }
    for attempt in range(1, 4):
        r = requests.post(
            f"{BASE_URL}/agent_config/upsert_agent_config", json=payload, headers=HEADERS
        )
        if r.status_code == 500 and attempt < 3:
            print(f"    agent {agent_id}: 500 error, retrying ({attempt}/3)...")
            time.sleep(2)
            continue
        r.raise_for_status()
        break


def poll_completed_calls(agent_ids: list[int], batch_start_ts: int) -> dict[int, dict]:
    """Return mapping of agent_id -> {id, metrics, to_number, transcript} for completed calls."""
    deadline = time.time() + POLL_TIMEOUT_S
    found: dict[int, dict] = {}
    conn = psycopg2.connect(DB_URL)
    try:
        while time.time() < deadline:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT DISTINCT ON (agent_id)
                        agent_id, id, metrics, to_number, transcript
                    FROM call_logs
                    WHERE agent_id = ANY(%s)
                      AND created_at >= %s
                      AND status = 'completed'
                    ORDER BY agent_id, created_at DESC
                    """,
                    (agent_ids, batch_start_ts),
                )
                for agent_id, call_log_id, metrics, to_number, transcript in cur.fetchall():
                    found[agent_id] = {
                        "id": call_log_id,
                        "metrics": metrics,
                        "to_number": to_number,
                        "transcript": transcript,
                    }
            conn.commit()
            missing = [a for a in agent_ids if a not in found]
            print(
                f"  poll: {len(found)}/{len(agent_ids)} completed"
                + (f" (waiting on {missing})" if missing else "")
            )
            if not missing:
                return found
            if time.time() >= deadline:
                print(f"  timeout reached — {len(found)} completed, {len(missing)} not completed")
                return found
            time.sleep(POLL_INTERVAL_S)
    finally:
        conn.close()
    return found


def main() -> None:
    agents = load_agents()
    agent_ids = [a["agent_id"] for a in agents]
    print(f"Loaded {len(agents)} agents: {agent_ids}")

    # Pre-fetch current configs so we don't lose system_prompt/first_message.
    base_configs = {a["agent_id"]: fetch_existing_config(a["agent_id"]) for a in agents}

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["combinations"]

    # Collect row indexes needing processing (status != 'completed')
    pending_rows = [
        r for r in range(2, ws.max_row + 1)
        if (ws.cell(row=r, column=COL["status"]).value or "").lower() != "completed"
    ]
    print(f"{len(pending_rows)} pending rows in {EXCEL_PATH.name}")

    for batch_start in range(0, len(pending_rows), BATCH_SIZE):
        batch = pending_rows[batch_start : batch_start + BATCH_SIZE]
        if len(batch) < BATCH_SIZE:
            print(f"Skipping final partial batch of {len(batch)} rows")
            break

        print(f"\n=== Batch {batch_start // BATCH_SIZE + 1}: rows {batch} ===")

        # 1. Update agent configs
        for agent, row_idx in zip(agents, batch):
            agent_id = agent["agent_id"]
            row_data = {
                "llm_sp_id": ws.cell(row=row_idx, column=COL["llm_sp_id"]).value,
                "llm_model_id": ws.cell(row=row_idx, column=COL["llm_model_id"]).value,
                "stt_sp_id": ws.cell(row=row_idx, column=COL["stt_sp_id"]).value,
                "stt_model_id": ws.cell(row=row_idx, column=COL["stt_model_id"]).value,
                "tts_sp_id": ws.cell(row=row_idx, column=COL["tts_sp_id"]).value,
                "tts_model_id": ws.cell(row=row_idx, column=COL["tts_model_id"]).value,
            }
            try:
                update_agent_config(agent_id, row_data, base_configs[agent_id])
                print(
                    f"  agent {agent_id} ({agent['phone_number']}) <- row {row_idx}: "
                    f"llm={row_data['llm_sp_id']}/{row_data['llm_model_id']} "
                    f"stt={row_data['stt_sp_id']}/{row_data['stt_model_id']} "
                    f"tts={row_data['tts_sp_id']}/{row_data['tts_model_id']}"
                )
            except requests.HTTPError as e:
                print(f"  FAILED to update agent {agent_id}: {e} | {e.response.text}")
                raise

        # 1b. Verify configs were applied correctly; retry if any service IDs are missing
        MAX_VERIFY_RETRIES = 3
        for attempt in range(1, MAX_VERIFY_RETRIES + 1):
            mismatched = []
            for agent, row_idx in zip(agents, batch):
                agent_id = agent["agent_id"]
                current = fetch_existing_config(agent_id)
                expected_llm = ws.cell(row=row_idx, column=COL["llm_sp_id"]).value
                expected_stt = ws.cell(row=row_idx, column=COL["stt_sp_id"]).value
                expected_tts = ws.cell(row=row_idx, column=COL["tts_sp_id"]).value
                cur_llm = current.get("llm_service_id")
                cur_stt = current.get("stt_service_id")
                cur_tts = current.get("tts_service_id")
                if cur_llm != expected_llm or cur_stt != expected_stt or cur_tts != expected_tts:
                    mismatched.append((agent, row_idx, {
                        "llm": f"{cur_llm} (expected {expected_llm})",
                        "stt": f"{cur_stt} (expected {expected_stt})",
                        "tts": f"{cur_tts} (expected {expected_tts})",
                    }))

            if not mismatched:
                print(f"  All 10 agent configs verified OK")
                break

            print(f"  Verify attempt {attempt}/{MAX_VERIFY_RETRIES}: {len(mismatched)} agents have mismatched configs")
            for agent, row_idx, diff in mismatched:
                print(f"    agent {agent['agent_id']} row {row_idx}: {diff}")
                row_data = {
                    "llm_sp_id": ws.cell(row=row_idx, column=COL["llm_sp_id"]).value,
                    "llm_model_id": ws.cell(row=row_idx, column=COL["llm_model_id"]).value,
                    "stt_sp_id": ws.cell(row=row_idx, column=COL["stt_sp_id"]).value,
                    "stt_model_id": ws.cell(row=row_idx, column=COL["stt_model_id"]).value,
                    "tts_sp_id": ws.cell(row=row_idx, column=COL["tts_sp_id"]).value,
                    "tts_model_id": ws.cell(row=row_idx, column=COL["tts_model_id"]).value,
                }
                try:
                    update_agent_config(agent["agent_id"], row_data, base_configs[agent["agent_id"]])
                    print(f"    -> retried update for agent {agent['agent_id']}")
                except requests.HTTPError as e:
                    print(f"    -> retry FAILED: {e} | {e.response.text}")
            time.sleep(2)
        else:
            print(f"  WARNING: {len(mismatched)} agents still have mismatched configs after {MAX_VERIFY_RETRIES} retries")

        # 1c. Flush Redis cache for all agents so stale bot data isn't used
        flush_redis_cache(agent_ids)

        # 2. Wait for user to trigger the calls
        batch_start_ts = int(time.time())
        input(
            f"\nConfigs updated & verified. Place the 10 calls now, then press Enter to begin "
            f"polling call_logs (batch start ts={batch_start_ts})... "
        )

        # 3. Poll
        results = poll_completed_calls(agent_ids, batch_start_ts)

        # 4. Write results back
        agent_to_row = {a["agent_id"]: r for a, r in zip(agents, batch)}
        for agent_id, row_idx in agent_to_row.items():
            if agent_id in results:
                res = results[agent_id]
                ws.cell(row=row_idx, column=COL["call_log_id"]).value = res["id"]
                ws.cell(row=row_idx, column=COL["status"]).value = "completed"
                ws.cell(row=row_idx, column=COL["metrics"]).value = (
                    json.dumps(res["metrics"]) if res["metrics"] is not None else None
                )
                ws.cell(row=row_idx, column=COL["transcription"]).value = (
                    json.dumps(res["transcript"]) if res["transcript"] is not None else None
                )
                ws.cell(row=row_idx, column=COL["to_number"]).value = res["to_number"]
            else:
                ws.cell(row=row_idx, column=COL["status"]).value = "not_completed"
                print(f"  WARNING: agent {agent_id} row {row_idx} not completed")

        wb.save(EXCEL_PATH)
        print(f"Saved batch results to {EXCEL_PATH}")

    print("\nAll batches done.")


if __name__ == "__main__":
    main()
