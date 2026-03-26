"""
Automated TTS voice testing pipeline.

Flow (repeats until all Excel rows are processed):
1. Pick next 10 unprocessed rows from the Excel file.
2. Update each agent's TTS config (tts_service_id + tts_metadata) via the API.
3. Wait for the user to trigger calls from the external testing app.
4. Poll the DB every 60 seconds until all 10 calls complete.
5. Check transcripts → mark each row as "pass" or "fail" in the Excel.

Usage:
    python scripts/update_tts_test_status.py
"""

import sys
import os
import time
import requests
import openpyxl
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------------------
# Configuration — fill these in
# ---------------------------------------------------------------------------

EXCEL_PATH = "/Users/thilak/Documents/Tone/scripts/TTS_Test_Configurations.xlsx"
DATABASE_URL = "postgresql://neondb_owner:npg_iNWhZLF0gHt7@ep-holy-wind-ad79pdco-pooler.c-2.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"

API_BASE_URL = "http://localhost:8000/api/v1"
JWT_TOKEN = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyX2lkIjoxLCJlbWFpbCI6InRoaWxhay5ndW5hc2VrYXJhbkBwcm9kdWN0ZnVzaW9uLmNvIiwib3JnX2lkIjoiYTc5MDUyZWEtNWFlYS00NzRkLWJlY2MtNTU4ZjY2ZWQ5ZGU0Iiwicm9sZSI6Im93bmVyIiwiaWF0IjoxNzc0MzQ1NjgyLCJleHAiOjE3NzQ0MzIwODJ9.4Jv4P4MnthtOQMH-CFK2nUOztWukvfrCjJxdsRrn0ng"

# 10 agent IDs — one per concurrent test slot
AGENT_IDS = [52, 248, 10891, 10892, 10893, 10894, 10895, 10899, 10897, 10898]
# AGENT_IDS = [248, 10892, 10893, 10894, 52]

POLL_INTERVAL_SECONDS = 60
BATCH_SIZE = 10

# ---------------------------------------------------------------------------
# SQL queries
# ---------------------------------------------------------------------------

# Look up service_provider DB id from name
PROVIDER_ID_QUERY = text("""
    SELECT id FROM service_providers
    WHERE name = :name
    LIMIT 1
""")

# Look up model DB id from name + provider
MODEL_ID_QUERY = text("""
    SELECT id FROM models
    WHERE name = :name AND service_provider_id = :provider_id AND service_type = 'tts'
    LIMIT 1
""")

# Get existing agent config (need system_prompt for the upsert API)
AGENT_CONFIG_QUERY = text("""
    SELECT id, uuid, system_prompt, tts_service_id, tts_metadata
    FROM agent_configs
    WHERE agent_id = :agent_id
    LIMIT 1
""")

# Get full agent config for reference agent (agent_id=52) to preserve all fields
REFERENCE_AGENT_CONFIG_QUERY = text("""
    SELECT id, uuid, system_prompt, llm_service_id, tts_service_id, stt_service_id,
           first_message, end_call_message, voicemail_message, html_prompt, status,
           llm_metadata, tts_metadata, stt_metadata, agent_metadata, description
    FROM agent_configs
    WHERE agent_id = :agent_id
    LIMIT 1
""")

# Check for completed calls after a given timestamp
CALL_COMPLETE_QUERY = text("""
    SELECT agent_id, transcript, id, to_number
    FROM call_logs
    WHERE agent_id = ANY(:agent_ids)
      AND started_at >= :since
      AND (status != 'in_progress' OR ended_at IS NOT NULL)
    ORDER BY started_at DESC
""")

# ANSI colors
GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
CYAN = "\033[96m"
BOLD = "\033[1m"
RESET = "\033[0m"


def get_unprocessed_rows(ws, status_col):
    """Return list of (row_index,) for rows where status is empty/None."""
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=status_col).value
        if val is None or str(val).strip() == "":
            rows.append(row_idx)
    return rows


def resolve_provider_id(conn, provider_name):
    """Get the DB integer ID for a service provider by name."""
    result = conn.execute(PROVIDER_ID_QUERY, {"name": provider_name})
    row = result.fetchone()
    return row[0] if row else None


def resolve_model_id(conn, model_name, provider_id):
    """Get the DB integer ID for a model by name + provider."""
    result = conn.execute(MODEL_ID_QUERY, {"name": model_name, "provider_id": provider_id})
    row = result.fetchone()
    return row[0] if row else None


def get_agent_config(conn, agent_id):
    """Fetch existing agent config from DB."""
    result = conn.execute(AGENT_CONFIG_QUERY, {"agent_id": agent_id})
    return result.fetchone()


def get_reference_agent_config(conn, agent_id=52):
    """Fetch full config from the reference agent to use as base for all upserts."""
    result = conn.execute(REFERENCE_AGENT_CONFIG_QUERY, {"agent_id": agent_id})
    row = result.fetchone()
    if not row:
        return None
    return {
        "system_prompt": row[2] or "You are a helpful voice assistant.",
        "llm_service_id": row[3],
        "stt_service_id": row[5],
        "first_message": row[6],
        "end_call_message": row[7],
        "voicemail_message": row[8],
        "html_prompt": row[9],
        "status": row[10] or "active",
        "llm_metadata": row[11],
        "stt_metadata": row[13],
        "agent_metadata": row[14],
        "description": row[15],
    }


def update_agent_config_via_api(agent_id, config_id, ref_config, tts_service_id, tts_metadata):
    """Call the upsert_agent_config API to update TTS settings.

    Uses ref_config (from reference agent 52) as the base payload so that
    llm_metadata, stt_metadata, and other fields are preserved. Only
    tts_service_id and tts_metadata are overridden per Excel row.
    """
    url = f"{API_BASE_URL}/agent_config/upsert_agent_config"
    headers = {
        "Authorization": f"Bearer {JWT_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "agent_id": agent_id,
        "id": config_id,
        **ref_config,
        "tts_service_id": tts_service_id,
        "tts_metadata": tts_metadata,
    }
    resp = requests.post(url, json=payload, headers=headers)
    if resp.status_code != 200:
        print(f"  {RED}API error for agent {agent_id}: {resp.status_code} {resp.text}{RESET}")
        return False
    return True


def poll_for_completed_calls(conn, agent_ids, since_ts, agent_info=None):
    """
    Poll DB until all agents have a completed call_log entry.
    Returns dict {agent_id: {transcript, call_log_id, to_number}}.

    agent_info: optional dict {agent_id: {provider, model, voice}} for display.
    """
    pending = set(agent_ids)
    results = {}

    while pending:
        result = conn.execute(CALL_COMPLETE_QUERY, {
            "agent_ids": list(pending),
            "since": since_ts,
        })
        newly_completed = []
        for row in result:
            aid = row[0]
            if aid in pending:
                call_data = {
                    "transcript": row[1],
                    "call_log_id": row[2],
                    "to_number": row[3],
                }
                results[aid] = call_data
                pending.discard(aid)
                newly_completed.append((aid, call_data))

        # Print details for each agent that just completed
        for aid, call_data in newly_completed:
            status = "pass" if has_valid_transcript(call_data["transcript"]) else "fail"
            color = GREEN if status == "pass" else RED
            info = agent_info.get(aid, {}) if agent_info else {}
            combo = f"{info.get('provider', '?')} / {info.get('model', '?')} / {info.get('voice', '?')}"
            print(f"  {color}✓ Agent {aid}: {status.upper()} — {combo} | "
                  f"call_log_id={call_data['call_log_id']} | "
                  f"to_number={call_data['to_number']}{RESET}")

        if pending:
            pending_list = ", ".join(str(a) for a in sorted(pending))
            print(f"  {YELLOW}Waiting... {len(pending)}/{len(agent_ids)} calls still pending. "
                  f"Agents: [{pending_list}]. "
                  f"Polling again in {POLL_INTERVAL_SECONDS}s{RESET}")
            time.sleep(POLL_INTERVAL_SECONDS)

    return results


def has_valid_transcript(transcript):
    """Check if a transcript is non-empty and meaningful."""
    if transcript is None:
        return False
    t = str(transcript).strip()
    return t not in ("", "null", "[]", "{}", "None")


def main():
    if not AGENT_IDS:
        print(f"{RED}Error: AGENT_IDS is empty. Add your 10 agent IDs.{RESET}")
        return
    if not API_BASE_URL:
        print(f"{RED}Error: API_BASE_URL is not set.{RESET}")
        return
    if not JWT_TOKEN:
        print(f"{RED}Error: JWT_TOKEN is not set.{RESET}")
        return

    engine = create_engine(DATABASE_URL)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    provider_col = headers["service_provider_id"]
    model_col = headers["model_name"]
    voice_col = headers["voice_id"]
    status_col = headers["status"]

    # Fetch reference agent config (agent_id=52) once — used as base for all upserts
    with engine.connect() as conn:
        ref_config = get_reference_agent_config(conn, agent_id=52)
    if not ref_config:
        print(f"{RED}Error: No agent_config found for reference agent_id=52.{RESET}")
        wb.close()
        return
    print(f"  {GREEN}Loaded reference config from agent_id=52{RESET}")

    total_rows = ws.max_row - 1
    batch_num = 0

    while True:
        unprocessed = get_unprocessed_rows(ws, status_col)
        if not unprocessed:
            print(f"\n{GREEN}{BOLD}All {total_rows} rows processed!{RESET}")
            break

        batch = unprocessed[:BATCH_SIZE]
        batch_num += 1
        print(f"\n{'='*60}")
        print(f"{BOLD}Batch {batch_num} — processing rows: {batch[0]}-{batch[-1]} "
              f"({len(batch)} rows, {len(unprocessed)} remaining){RESET}")
        print(f"{'='*60}")

        # --- Step 1: Update agent configs ---
        print(f"\n{CYAN}Step 1: Updating agent TTS configs via API...{RESET}")

        agent_row_map = {}  # {agent_id: row_idx}
        agent_info = {}  # {agent_id: {provider, model, voice}} for poll display

        with engine.connect() as conn:
            for i, row_idx in enumerate(batch):
                agent_id = AGENT_IDS[i]
                provider_name_raw = ws.cell(row=row_idx, column=provider_col).value
                provider_name = ws.cell(row=row_idx, column=headers["service_provider_name"]).value
                model_name = ws.cell(row=row_idx, column=model_col).value
                voice_id = ws.cell(row=row_idx, column=voice_col).value

                # Resolve provider name to DB id
                provider_db_id = resolve_provider_id(conn, provider_name_raw)
                if not provider_db_id:
                    print(f"  {RED}Row {row_idx}: No service_provider found for '{provider_name_raw}'. Marking fail.{RESET}")
                    ws.cell(row=row_idx, column=status_col).value = "fail"
                    continue

                model_db_id = resolve_model_id(conn, model_name, provider_db_id)

                # Get existing config for id
                config = get_agent_config(conn, agent_id)
                if not config:
                    print(f"  {RED}Row {row_idx}: No agent_config for agent {agent_id}. Marking fail.{RESET}")
                    ws.cell(row=row_idx, column=status_col).value = "fail"
                    continue
                config_id = config[0]

                tts_metadata = {"voice_id": voice_id}
                if model_db_id:
                    tts_metadata["model_id"] = model_db_id

                success = update_agent_config_via_api(
                    agent_id=agent_id,
                    config_id=config_id,
                    ref_config=ref_config,
                    tts_service_id=int(provider_db_id),
                    tts_metadata=tts_metadata,
                )

                if success:
                    agent_row_map[agent_id] = row_idx
                    voice_name = ws.cell(row=row_idx, column=headers["voice_name"]).value
                    agent_info[agent_id] = {
                        "provider": provider_name,
                        "model": model_name,
                        "voice": voice_name,
                    }
                    print(f"  {GREEN}Agent {agent_id} → {provider_name} / {model_name} / {voice_name}{RESET}")
                else:
                    ws.cell(row=row_idx, column=status_col).value = "fail"

        if not agent_row_map:
            print(f"  {RED}No agents configured in this batch. Skipping.{RESET}")
            wb.save(EXCEL_PATH)
            continue

        # --- Step 2: Wait for user to trigger calls ---
        print(f"\n{CYAN}Step 2: Trigger calls from your testing app for {len(agent_row_map)} agent(s).{RESET}")
        since_ts = int(time.time())
        input(f"{BOLD}Press Enter once calls have been triggered...{RESET}")

        # --- Step 3: Poll for call completion ---
        print(f"\n{CYAN}Step 3: Polling DB for call completion...{RESET}")

        with engine.connect() as conn:
            call_results = poll_for_completed_calls(
                conn,
                list(agent_row_map.keys()),
                since_ts,
                agent_info=agent_info,
            )

        # --- Step 4: Update Excel with results ---
        print(f"\n{CYAN}Step 4: Updating Excel...{RESET}")

        call_log_id_col = headers["call_log_id"]
        to_number_col = headers["to_number"]

        pass_count = 0
        fail_count = 0
        for agent_id, row_idx in agent_row_map.items():
            call_data = call_results.get(agent_id, {})
            transcript = call_data.get("transcript") if call_data else None
            call_log_id = call_data.get("call_log_id") if call_data else None
            to_number = call_data.get("to_number") if call_data else None

            ws.cell(row=row_idx, column=call_log_id_col).value = call_log_id
            ws.cell(row=row_idx, column=to_number_col).value = to_number

            if has_valid_transcript(transcript):
                ws.cell(row=row_idx, column=status_col).value = "pass"
                pass_count += 1
                print(f"  {GREEN}Row {row_idx}: pass (call_log_id={call_log_id}){RESET}")
            else:
                ws.cell(row=row_idx, column=status_col).value = "fail"
                fail_count += 1
                print(f"  {RED}Row {row_idx}: fail (call_log_id={call_log_id}){RESET}")

        wb.save(EXCEL_PATH)
        print(f"\n  Batch {batch_num} done — {GREEN}Pass: {pass_count}{RESET}, {RED}Fail: {fail_count}{RESET}")
        print(f"  Excel saved.")

    wb.close()
    print(f"\n{GREEN}{BOLD}Pipeline complete!{RESET}")


if __name__ == "__main__":
    main()
