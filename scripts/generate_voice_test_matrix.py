"""
Generate Voice_Test_Matrix.xlsx from dev-data.json for the 14 target TTS providers.

1:1 model-voice pairing logic:
- Most providers: all voices work with all models → pair sequentially
- Groq: language-specific models → match English/Arabic model to matching voice
- Rime: voices are model-specific (mist=English, mistv2=4 langs, arcana=9 langs)
- Sarvam: speakers are model-specific (bulbul:v2=7 speakers, bulbul:v3=39 speakers)

Usage:
    python scripts/generate_voice_test_matrix.py
"""

import json
import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


TARGET_PROVIDERS = {
    "Cartesia", "Groq", "Hume AI", "MiniMax", "Rime", "Sarvam",
    "Inworld", "OpenAI", "Neuphonic", "LMNT", "ElevenLabs",
    "Camb AI", "Async AI", "Speechmatics",
}

# ── Groq: language-specific models ──
GROQ_MODEL_LANGUAGE_MAP = {
    "canopylabs/orpheus-v1-english": "en",
    "canopylabs/orpheus-arabic-saudi": "ar",
}

# ── Rime: voices are model-specific (from Rime_Voices_Model_Language_Reference.md) ──
# mist: 117 English-only voices
RIME_MIST_VOICES = set(
    "abbie, alexis, allison, ally, alona, alpine, amber, ana, antoine, armon, "
    "bayou, benjamin, blaze, blossom, boulder, breeze, brenda, brittany, brook, "
    "carol, cedar, colin, courtney, cove, creek, dew, elena, elliot, ember, eva, "
    "falcon, fjord, flower, frank, gabriela, geoff, gerald, glacier, granite, grove, "
    "gulch, gypsum, hank, hawk, helen, hera, iris, ironwood, jen, joe, joy, juan, "
    "jungle, kendra, kendrick, kenneth, kevin, kris, lagoon, linda, loquat, lotus, "
    "madison, marge, marina, marissa, marsh, marta, maya, mesa, moon, moraine, "
    "nicholas, nyles, peak, pearl, petal, phil, rain, rainforest, reba, rex, rick, "
    "ritu, river, rob, rodney, rohan, rosco, samantha, sandy, selena, seth, sharon, "
    "spore, stan, steppe, stone, storm, stream, summit, talon, tamra, tanya, thunder, "
    "tibur, tj, tundra, tyler, violet, viv, wildflower, willow, wolf, yadira, zest, zion"
    .split(", ")
)

# mistv2: 141 voices (English + Spanish + French + German)
RIME_MISTV2_VOICES = set(
    "abbie, allison, ally, alona, alpine, amber, ana, antoine, armon, astra, bayou, "
    "blaze, blossom, boulder, breeze, brenda, brittany, brook, carol, cedar, colin, "
    "courtney, cove, cove_extra, creek, dew, elena, elliot, ember, eucalyptus, eva, "
    "falcon, fjord, flower, geoff, gerald, glacier, granite, grove, gulch, gypsum, "
    "hank, hawk, helen, hera, iris, ironwood, jen, joe, jose, joy, juan, jungle, "
    "karst, kendra, kendrick, kenneth, kevin, kris, lagoon, linda, loquat, lotus, "
    "madison, marge, mari, marina, marissa, marlu, marsh, marta, maya, mesa, mesa_extra, "
    "moon, moraine, nicholas, nyles, pablo, peak, pearl, petal, phil, rain, rainforest, "
    "reba, rex, rick, ritu, river, rob, rodney, rohan, rosco, runton, samantha, sandy, "
    "selena, seth, sharon, spore, stan, steppe, stone, storm, stream, summit, talon, "
    "tamra, tanya, thunder, tibur, tj, tundra, tyler, violet, viv, wildflower, willow, "
    "wolf, yadira, zest, zion, "
    "diego, dolores, isa, jose, lucia, mari, mateo, pablo, sofia, "
    "alois, juliette, marguerite, simone, "
    "amalia, frieda, karolina, klaus, maximilian"
    .split(", ")
)

# arcana: 269 voices (9 languages) — all remaining voices belong to arcana

# ── Sarvam: speakers are model-specific (from Sarvam_Voices_Model_Reference.md) ──
SARVAM_V2_SPEAKERS = {
    "anushka", "manisha", "vidya", "arya", "abhilash", "karun", "hitesh",
}
SARVAM_V3_SPEAKERS = {
    "shubh", "aditya", "rahul", "rohan", "amit", "dev", "ratan", "varun",
    "manan", "sumit", "kabir", "aayan", "ashutosh", "advait", "anand",
    "tarun", "sunny", "mani", "gokul", "vijay", "mohit", "rehan", "soham",
    "ritu", "priya", "neha", "pooja", "simran", "kavya", "ishita", "shreya",
    "roopa", "amelia", "sophia", "tanya", "shruti", "suhani", "kavitha", "rupali",
}


def load_dev_data():
    data_path = os.path.join(os.path.dirname(__file__), "..", "dev", "dev-data.json")
    with open(data_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _get_rime_model_for_voice(voice_id):
    """Determine which Rime model a voice belongs to.
    Priority: arcana-exclusive > mistv2-exclusive > mist-exclusive.
    Voices in multiple models are assigned to the most capable one."""
    in_mist = voice_id in RIME_MIST_VOICES
    in_mistv2 = voice_id in RIME_MISTV2_VOICES
    # If not in mist or mistv2, it's arcana-exclusive
    if not in_mist and not in_mistv2:
        return "arcana"
    # If in mistv2 but not mist, it's mistv2 (or shared with arcana)
    if in_mistv2 and not in_mist:
        return "mistv2"
    # If in mist but not mistv2, it's mist-exclusive
    if in_mist and not in_mistv2:
        return "mist"
    # In both mist and mistv2 — assign to mistv2 (more capable)
    return "mistv2"


def _get_sarvam_model_for_voice(voice_id):
    """Determine which Sarvam model a voice belongs to.
    voice_id format: sarvam-{speaker}-{lang}"""
    parts = voice_id.split("-")
    if len(parts) >= 2:
        speaker = parts[1].lower()
    else:
        speaker = voice_id.lower()
    if speaker in SARVAM_V2_SPEAKERS:
        return "bulbul:v2"
    if speaker in SARVAM_V3_SPEAKERS:
        return "bulbul:v3"
    return None


def pair_groq(models, voices):
    """Groq: match each model to a voice with the same language."""
    pairs = []
    used_voices = set()
    for model in models:
        lang = GROQ_MODEL_LANGUAGE_MAP.get(model["name"])
        if not lang:
            continue
        for v in voices:
            if v.get("voice_id") in used_voices:
                continue
            if lang in v.get("language_list", []):
                pairs.append((model, v))
                used_voices.add(v["voice_id"])
                break
    return pairs


def pair_rime(models, voices):
    """Rime: voices are model-specific. Pick one voice per model."""
    model_map = {m["name"]: m for m in models}
    # Group voices by their model
    voices_by_model = {"mist": [], "mistv2": [], "arcana": []}
    for v in voices:
        model_name = _get_rime_model_for_voice(v["voice_id"])
        if model_name in voices_by_model:
            voices_by_model[model_name].append(v)

    pairs = []
    for model_name in ["arcana", "mistv2", "mist"]:
        model = model_map.get(model_name)
        available = voices_by_model.get(model_name, [])
        if model and available:
            pairs.append((model, available[0]))
    return pairs


def pair_sarvam(models, voices):
    """Sarvam: speakers are model-specific. Pick one voice per model."""
    model_map = {m["name"]: m for m in models}
    # Group voices by their model
    voices_by_model = {"bulbul:v2": [], "bulbul:v3": []}
    for v in voices:
        model_name = _get_sarvam_model_for_voice(v["voice_id"])
        if model_name and model_name in voices_by_model:
            voices_by_model[model_name].append(v)

    pairs = []
    for model_name in ["bulbul:v3", "bulbul:v2"]:
        model = model_map.get(model_name)
        available = voices_by_model.get(model_name, [])
        if model and available:
            pairs.append((model, available[0]))
    return pairs


def pair_generic(models, voices):
    """Default: 1:1 sequential pairing."""
    pairs = []
    count = min(len(models), len(voices))
    for i in range(count):
        pairs.append((models[i], voices[i]))
    return pairs


def generate():
    data = load_dev_data()

    rows = []

    for provider in data.get("tts_providers", []):
        display_name = provider.get("display_name", "")
        if display_name not in TARGET_PROVIDERS:
            continue

        provider_name = provider["name"]
        models = provider.get("models", [])
        voices = provider.get("voices", [])

        if not models or not voices:
            print(f"  {display_name}: skipped (no models or voices)")
            continue

        # Pick pairing strategy
        if display_name == "Groq":
            pairs = pair_groq(models, voices)
        elif display_name == "Rime":
            pairs = pair_rime(models, voices)
        elif display_name == "Sarvam":
            pairs = pair_sarvam(models, voices)
        else:
            pairs = pair_generic(models, voices)

        print(f"  {display_name}: {len(models)} model(s), {len(voices)} voice(s) → {len(pairs)} pair(s)")

        for model, voice in pairs:
            rows.append({
                "service_provider_id": provider_name,
                "service_provider_name": display_name,
                "model_id": "",
                "model_name": model["name"],
                "voice_id": voice.get("voice_id", ""),
                "voice_name": voice.get("name", ""),
                "status": "",
                "call_log_id": "",
                "to_number": "",
            })

    if not rows:
        print("No pairs generated.")
        return

    # ── Create Excel workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voice Test Matrix"

    headers = [
        "service_provider_id",
        "service_provider_name",
        "model_id",
        "model_name",
        "voice_id",
        "voice_name",
        "status",
        "call_log_id",
        "to_number",
    ]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths
    total_rows = len(rows) + 1
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col - 1])
        for r in range(2, total_rows + 1):
            cell_value = ws.cell(row=r, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_length + 4, 50)

    # Freeze header row and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{total_rows}"

    # Save
    output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Voice_Test_Matrix.xlsx")
    wb.save(output_file)

    print(f"\nExcel file created: {output_file}")
    print(f"Total test pairs: {len(rows)}")


if __name__ == "__main__":
    generate()
