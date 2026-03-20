import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Target providers
TARGET_PROVIDERS = {
    "Cartesia", "Groq", "Hathora", "Minimax", "Rime", "Sarvam",
    "Inworld", "OpenAI", "Neuphonic", "LMNT", "Hume", "ElevenLabs",
    "Camb", "AsyncAI HTTP", "Speechmatics"
}

with open("dev/dev-data.json", "r") as f:
    data = json.load(f)

tts_providers = data.get("tts_providers", [])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TTS Test Configurations"

# Headers
headers = [
    "service_provider_name",
    "service_provider_id",
    "model_id",
    "model_name",
    "voice_id",
    "voice_name",
    "status",
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

row = 2
for provider in tts_providers:
    display_name = provider.get("display_name", provider.get("name", ""))
    provider_name = provider.get("name", "")

    if display_name not in TARGET_PROVIDERS:
        continue

    models = provider.get("models", [])
    voices = provider.get("voices", [])

    if not models:
        models = [{"name": "default", "meta_data": {}}]

    if not voices:
        voices = [{"name": "N/A", "voice_id": "N/A"}]

    for model in models:
        model_name = model.get("name", "")
        for voice in voices:
            voice_name = voice.get("name", "")
            voice_id = voice.get("voice_id", "")

            ws.cell(row=row, column=1, value=display_name).border = thin_border
            ws.cell(row=row, column=2, value=provider_name).border = thin_border
            ws.cell(row=row, column=3, value="").border = thin_border  # model_id from DB, left empty
            ws.cell(row=row, column=4, value=model_name).border = thin_border
            ws.cell(row=row, column=5, value=voice_id).border = thin_border
            ws.cell(row=row, column=6, value=voice_name).border = thin_border
            ws.cell(row=row, column=7, value="").border = thin_border  # status - empty

            row += 1

# Auto-fit column widths
for col in range(1, len(headers) + 1):
    max_length = len(headers[col - 1])
    for r in range(2, row):
        cell_value = ws.cell(row=r, column=col).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_length + 4, 50)

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:G{row - 1}"

output_file = "TTS_Test_Configurations.xlsx"
wb.save(output_file)
print(f"Excel file created: {output_file}")
print(f"Total rows (excluding header): {row - 2}")
