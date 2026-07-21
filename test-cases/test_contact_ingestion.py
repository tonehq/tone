"""Unit tests for contact ingestion + per-row schedule-time resolution.

Pure logic, no DB:
- core/services/contact_ingestion/csv_source.py — CSV parsing/normalization.
- core/services/contact_ingestion/excel_source.py — .xlsx parsing/coercion.
- core/services/contact_ingestion/row_mapping.py — the shared header→ParsedContact mapper.
- core/services/contact_ingestion/__init__.py — source factory + upload selector.
- core/services/outbound_call_service.py::_resolve_contact_when — per-row schedule time.
"""

import io
from datetime import datetime, timezone

import openpyxl
import pytest

from core.services.contact_ingestion import (
    get_contact_source,
    select_source_for_upload,
)
from core.services.contact_ingestion.base import ParsedContact
from core.services.contact_ingestion.validation import (
    PhoneNumberValidator,
    RequiredIdentityValidator,
    SchemaMetadataValidator,
    build_contact_validator,
)
from core.services.contact_ingestion.csv_source import CSVContactSource
from core.services.contact_ingestion.excel_source import ExcelContactSource
from core.services.outbound_call_service import OutboundCallService as Svc


def _parse(raw: str):
    return list(CSVContactSource().parse(raw.encode("utf-8")))


def _xlsx_bytes(rows) -> bytes:
    """Build an in-memory ``.xlsx`` from ``rows`` (each a list of cell values)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_xlsx(rows):
    return list(ExcelContactSource().parse(_xlsx_bytes(rows)))


class TestCSVParsing:
    def test_one_parsed_contact_per_row(self):
        rows = _parse("name,phone_number\nAlice,+14155550123\nBob,+14155550124\n")
        assert len(rows) == 2
        assert rows[0].name == "Alice"
        assert rows[0].phone_number == "+14155550123"

    def test_headers_are_case_insensitive_and_aliased(self):
        rows = _parse("Name,Phone\nAlice,+14155550123\n")
        assert rows[0].name == "Alice"
        assert rows[0].phone_number == "+14155550123"

    def test_scheduled_at_goes_to_metadata_iso_utc(self):
        rows = _parse("name,phone,scheduled_at\nA,+14155550123,2026-08-01T15:30:00Z\n")
        assert rows[0].metadata["scheduled_at"] == "2026-08-01T15:30:00+00:00"

    def test_naive_scheduled_at_assumed_utc(self):
        rows = _parse("name,phone,call_time\nA,+14155550123,2026-08-01T15:30:00\n")
        assert rows[0].metadata["scheduled_at"] == "2026-08-01T15:30:00+00:00"

    def test_extra_columns_flow_to_metadata(self):
        rows = _parse("name,phone,company,notes\nA,+14155550123,Acme,VIP\n")
        assert rows[0].metadata["company"] == "Acme"
        assert rows[0].metadata["notes"] == "VIP"

    def test_required_marker_star_stripped_from_headers(self):
        # Sample templates mark required columns with a trailing ``*`` (e.g. ``city*``);
        # the marker is stripped so a filled-in template still maps to the real field.
        rows = _parse("name*,phone_number*,city*\nJohn Doe,+14155550123,SF\n")
        assert rows[0].name == "John Doe"
        assert rows[0].phone_number == "+14155550123"
        assert rows[0].metadata["city"] == "SF"

    def test_external_id_used_when_present(self):
        rows = _parse("external_id,name,phone\ncust-1,A,+14155550123\n")
        assert rows[0].external_id == "cust-1"

    def test_external_id_synthesized_from_phone_when_absent(self):
        rows = _parse("name,phone\nA,+14155550123\n")
        assert rows[0].external_id == "+14155550123"

    def test_external_id_stable_hash_when_no_phone(self):
        rows1 = _parse("name,company\nA,Acme\n")
        rows2 = _parse("name,company\nA,Acme\n")
        # Same row content → same synthesized id (idempotent re-import).
        assert rows1[0].external_id == rows2[0].external_id
        assert rows1[0].external_id.startswith("csv-")

    def test_blank_rows_skipped(self):
        rows = _parse("name,phone\nA,+14155550123\n,\n")
        assert len(rows) == 1

    def test_phone_normalization_strips_formatting(self):
        rows = _parse("name,phone\nA,(415) 555-0123\n")
        assert rows[0].phone_number == "4155550123"

    def test_phone_00_prefix_becomes_plus(self):
        rows = _parse("name,phone\nA,0044 20 7946 0000\n")
        assert rows[0].phone_number == "+442079460000"

    def test_unparseable_scheduled_at_dropped(self):
        rows = _parse("name,phone,scheduled_at\nA,+14155550123,not-a-date\n")
        assert "scheduled_at" not in rows[0].metadata

    def test_bom_tolerated(self):
        rows = list(CSVContactSource().parse("﻿name,phone\nA,+14155550123\n".encode("utf-8")))
        assert rows[0].name == "A"

    def test_ragged_row_with_trailing_comma_does_not_crash(self):
        # csv.DictReader collects overflow cells (a trailing comma / extra column) into a
        # list under the None restkey; the mapper must tolerate the non-string value instead
        # of raising AttributeError and aborting the whole import.
        rows = _parse("name,phone\nAlice,+14155550123,\nBob,+14155550124,extra\n")
        assert len(rows) == 2
        assert rows[0].name == "Alice"
        assert rows[0].phone_number == "+14155550123"
        assert rows[1].name == "Bob"

    def test_ragged_row_without_phone_hashes_without_crash(self):
        # A ragged row (extra trailing cell → None restkey) that ALSO lacks a phone and
        # external_id falls into the content-hash branch; sorting row.items() there must not
        # choke on the None key (would raise TypeError and abort the whole import).
        rows = _parse("name,city\nAlice,NYC,extra\n")
        assert len(rows) == 1
        assert rows[0].name == "Alice"
        assert rows[0].phone_number is None
        assert rows[0].external_id.startswith("csv-")


class _FakeDatasource:
    """Minimal stand-in for a ``Datasource`` row — the factory only reads ``.type``."""

    def __init__(self, type):
        self.type = type


class TestSourceFactory:
    def test_csv_default(self):
        assert isinstance(get_contact_source(_FakeDatasource("csv")), CSVContactSource)

    def test_csv_case_insensitive(self):
        assert isinstance(get_contact_source(_FakeDatasource("CSV")), CSVContactSource)

    def test_unknown_source_raises(self):
        with pytest.raises(ValueError):
            get_contact_source(_FakeDatasource("xlsx"))


class TestExcelParsing:
    def test_xlsx_matches_equivalent_csv(self):
        # A .xlsx and its equivalent CSV must produce identical ParsedContacts (same
        # mapper), proving CSV/Excel parity: same external_id, name, phone, metadata.
        csv_rows = _parse("name,phone_number,company\nAlice,+14155550123,Acme\n")
        xlsx_rows = _parse_xlsx(
            [["name", "phone_number", "company"], ["Alice", "+14155550123", "Acme"]]
        )
        assert len(xlsx_rows) == len(csv_rows) == 1
        c, x = csv_rows[0], xlsx_rows[0]
        assert (x.external_id, x.name, x.phone_number, x.metadata) == (
            c.external_id,
            c.name,
            c.phone_number,
            c.metadata,
        )

    def test_numeric_phone_cell_not_mangled(self):
        # A phone stored as a NUMBER in Excel must not gain a ``.0`` or scientific notation.
        rows = _parse_xlsx([["name", "phone_number"], ["Alice", 14155550123]])
        assert rows[0].phone_number == "14155550123"
        assert rows[0].external_id == "14155550123"

    def test_integral_float_cell_has_no_trailing_zero(self):
        rows = _parse_xlsx([["name", "age"], ["Alice", 42.0]])
        assert rows[0].metadata["age"] == "42"

    def test_datetime_cell_maps_to_utc_scheduled_at(self):
        # A datetime cell coerces to ISO and (naive) is assumed UTC — same as the CSV path.
        rows = _parse_xlsx(
            [["name", "phone", "scheduled_at"], ["A", "+14155550123", datetime(2026, 8, 1, 15, 30)]]
        )
        assert rows[0].metadata["scheduled_at"] == "2026-08-01T15:30:00+00:00"

    def test_leading_blank_rows_before_header_skipped(self):
        rows = _parse_xlsx(
            [
                [None, None],  # spacer/title row
                ["", ""],  # blank row
                ["name", "phone_number"],
                ["Alice", "+14155550123"],
            ]
        )
        assert len(rows) == 1
        assert rows[0].name == "Alice"

    def test_duplicate_headers_later_column_wins(self):
        # Mirror csv.DictReader: the later same-named column overrides the earlier one.
        rows = _parse_xlsx(
            [["name", "phone_number", "phone_number"], ["A", "+14155550111", "+14155550222"]]
        )
        assert rows[0].phone_number == "+14155550222"

    def test_only_active_sheet_is_read(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "phone_number"])
        ws.append(["Alice", "+14155550123"])
        other = wb.create_sheet("Ignored")
        other.append(["name", "phone_number"])
        other.append(["Bob", "+14155550999"])
        buf = io.BytesIO()
        wb.save(buf)
        rows = list(ExcelContactSource().parse(buf.getvalue()))
        assert [r.name for r in rows] == ["Alice"]

    def test_headers_only_yields_nothing(self):
        assert _parse_xlsx([["name", "phone_number"]]) == []

    def test_required_marker_star_stripped_in_xlsx(self):
        rows = _parse_xlsx([["name*", "phone_number*", "city*"], ["John", "+14155550123", "SF"]])
        assert rows[0].name == "John"
        assert rows[0].metadata["city"] == "SF"

    def test_corrupt_xlsx_raises(self):
        # PK magic but not a valid zip → openpyxl raises; the sync layer turns this into a
        # clean ``failed`` status (see contact_sync_service run wrapper).
        with pytest.raises(Exception):
            list(ExcelContactSource().parse(b"PK\x03\x04not-a-real-xlsx"))


class TestUploadSourceSelector:
    def test_xlsx_magic_selects_excel_source(self):
        raw = _xlsx_bytes([["name", "phone_number"], ["A", "+14155550123"]])
        assert isinstance(select_source_for_upload(raw), ExcelContactSource)

    def test_csv_bytes_default_to_csv_source(self):
        assert isinstance(
            select_source_for_upload(b"name,phone_number\nA,+14155550123\n"), CSVContactSource
        )

    def test_legacy_xls_ole2_magic_rejected(self):
        # OLE2 compound-file signature (legacy .xls) → friendly ValueError, not a parser.
        with pytest.raises(ValueError, match="Legacy .xls"):
            select_source_for_upload(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1rest-of-file")

    def test_empty_bytes_default_to_csv_source(self):
        assert isinstance(select_source_for_upload(b""), CSVContactSource)


class _FakeContact:
    def __init__(self, meta=None):
        self.id = "c-1"
        self.contact_metadata = meta or {}


class TestResolveContactWhen:
    def test_metadata_time_wins(self):
        future = "2999-01-01T00:00:00+00:00"
        c = _FakeContact({"scheduled_at": future})
        got = Svc._resolve_contact_when(c, None)
        assert got == datetime.fromisoformat(future)

    def test_request_time_used_when_no_metadata(self):
        req = datetime(2999, 6, 1, tzinfo=timezone.utc)
        got = Svc._resolve_contact_when(_FakeContact(), req)
        assert got == req

    def test_past_metadata_time_collapses_to_now(self):
        c = _FakeContact({"scheduled_at": "2000-01-01T00:00:00+00:00"})
        got = Svc._resolve_contact_when(c, None)
        assert (datetime.now(timezone.utc) - got).total_seconds() < 5

    def test_none_when_no_time_is_asap(self):
        got = Svc._resolve_contact_when(_FakeContact(), None)
        assert (datetime.now(timezone.utc) - got).total_seconds() < 5

    def test_unparseable_metadata_falls_back(self):
        req = datetime(2999, 6, 1, tzinfo=timezone.utc)
        c = _FakeContact({"scheduled_at": "garbage"})
        assert Svc._resolve_contact_when(c, req) == req


def test_parsed_contact_defaults():
    pc = ParsedContact(external_id="x")
    assert pc.name is None and pc.phone_number is None and pc.metadata == {}


# --------------------------------------------------------------------------- #
# parse_datetime_value — the shared datetime parser (scheduled_at + datetime fields)
# --------------------------------------------------------------------------- #

from types import SimpleNamespace  # noqa: E402

from core.services.contact_ingestion.contact_mapping import map_source_row_to_contact  # noqa: E402
from core.services.contact_ingestion.row_mapping import parse_datetime_value  # noqa: E402


class TestParseDatetimeValue:
    def test_blank_returns_none(self):
        assert parse_datetime_value("") is None
        assert parse_datetime_value("   ") is None

    def test_naive_iso_assumed_utc(self):
        assert parse_datetime_value("2026-07-28T10:30:00") == "2026-07-28T10:30:00+00:00"

    def test_iso_offset_preserved_and_normalized_to_utc(self):
        # -04:00 wall clock 10:30 → 14:30Z
        assert parse_datetime_value("2026-07-28T10:30:00-04:00") == "2026-07-28T14:30:00+00:00"

    def test_trailing_z_tolerated(self):
        assert parse_datetime_value("2026-07-28T10:30:00Z") == "2026-07-28T10:30:00+00:00"

    def test_custom_format_with_source_timezone(self):
        # Naive value parsed with an explicit format, interpreted in America/New_York (EDT,
        # -04:00) → normalized to UTC.
        got = parse_datetime_value(
            "07/28/2026 10:30", fmt="%m/%d/%Y %H:%M", tz="America/New_York"
        )
        assert got == "2026-07-28T14:30:00+00:00"

    def test_unknown_timezone_falls_back_to_utc(self):
        got = parse_datetime_value("2026-07-28T10:30:00", tz="Not/AZone")
        assert got == "2026-07-28T10:30:00+00:00"

    def test_unparseable_returns_none(self):
        assert parse_datetime_value("garbage") is None
        assert parse_datetime_value("2026-13-99", fmt="%Y-%m-%d") is None


class TestDatetimeSchemaFieldMapping:
    @staticmethod
    def _field(**kw):
        base = dict(
            source_key=None,
            field_name="call_time",
            type="string",
            format="datetime",
            field_metadata={},
        )
        base.update(kw)
        return SimpleNamespace(**base)

    def test_datetime_field_normalized_to_utc_iso(self):
        field = self._field(
            source_key="call_time",
            field_metadata={"datetime_format": "%m/%d/%Y %H:%M", "timezone": "America/New_York"},
        )
        out = map_source_row_to_contact({"call_time": "07/28/2026 10:30"}, [field])
        assert out["contact_metadata"]["call_time"] == "2026-07-28T14:30:00+00:00"

    def test_datetime_field_iso_fallback(self):
        field = self._field(source_key="call_time")
        out = map_source_row_to_contact({"call_time": "2026-07-28T10:30:00"}, [field])
        assert out["contact_metadata"]["call_time"] == "2026-07-28T10:30:00+00:00"

    def test_unparseable_datetime_dropped(self):
        field = self._field(source_key="call_time")
        out = map_source_row_to_contact({"call_time": "not a date"}, [field])
        assert "call_time" not in out["contact_metadata"]


# --------------------------------------------------------------------------- #
# Reusable Parser + Validator framework (pipeline.py + validation.py)
# --------------------------------------------------------------------------- #

from core.services.contact_ingestion.pipeline import (  # noqa: E402
    RecordParser,
    parsed_contact_to_row,
)
from core.services.contact_ingestion.validation import (  # noqa: E402
    CompositeValidator,
    PhoneNumberValidator,
    RecordValidator,
    RequiredIdentityValidator,
    SchemaMetadataValidator,
)


def _pc(external_id="x", name=None, phone=None, metadata=None):
    return ParsedContact(external_id=external_id, name=name, phone_number=phone, metadata=metadata or {})


def _schema_field(field_name, *, type="string", is_mandatory=False, format=None, validators=None):
    return SimpleNamespace(
        field_name=field_name,
        type=type,
        is_mandatory=is_mandatory,
        format=format,
        validators=validators or {},
        options=None,
    )


class TestValidators:
    def test_phone_valid(self):
        assert PhoneNumberValidator().validate(_pc(phone="+14155550123")) == []

    def test_phone_missing(self):
        assert PhoneNumberValidator().validate(_pc(phone=None)) == [
            "A phone number is required."
        ]

    def test_phone_not_e164(self):
        errs = PhoneNumberValidator().validate(_pc(phone="12345"))
        assert len(errs) == 1 and "E.164" in errs[0]

    def test_required_identity(self):
        assert RequiredIdentityValidator().validate(_pc(name="A")) == []
        assert RequiredIdentityValidator().validate(_pc(phone="+14155550123")) == []
        assert RequiredIdentityValidator().validate(_pc()) == [
            "A contact needs at least a name or phone number."
        ]

    def test_schema_metadata(self):
        v = SchemaMetadataValidator([_schema_field("city", is_mandatory=True)])
        assert v.validate(_pc(metadata={"city": "SF"})) == []
        # Missing the mandatory 'city' → an error.
        assert v.validate(_pc(metadata={})) != []

    def test_composite_aggregates_and_is_extensible(self):
        composite = CompositeValidator([PhoneNumberValidator()])
        # Empty composite validates everything.
        assert CompositeValidator().validate(_pc()) == []
        # A bad record fails the phone rule.
        assert composite.validate(_pc(phone=None))
        # Rules are added dynamically without touching the loop.
        composite.add(RequiredIdentityValidator())
        errs = composite.validate(_pc())  # no phone + no identity → 2 rules fire
        assert len(errs) == 2

    def test_custom_rule_plugs_in(self):
        class NoTestNumbers(RecordValidator):
            def validate(self, record):
                return ["Test number blocked."] if (record.phone_number or "").endswith("0000") else []

        composite = CompositeValidator([PhoneNumberValidator(), NoTestNumbers()])
        assert composite.validate(_pc(phone="+14155550123")) == []
        assert composite.validate(_pc(phone="+14155550000")) == ["Test number blocked."]


class TestRecordParser:
    def test_partitions_valid_and_invalid(self):
        records = [
            _pc(external_id="a", phone="+14155550123"),
            _pc(external_id="b", phone="bad"),
            _pc(external_id="c", phone="+14155550124"),
        ]
        result = RecordParser(CompositeValidator([PhoneNumberValidator()])).process(records)
        assert result.total == 3
        assert result.valid_count == 2 and result.invalid_count == 1
        assert result.invalid[0]["phone_number"] == "bad"
        assert [r.external_id for r in result.valid] == ["a", "c"]

    def test_truncation_reports_and_stops(self):
        records = [_pc(external_id=str(i), phone="+1415555010" + str(i)) for i in range(5)]
        result = RecordParser(CompositeValidator(), max_records=2).process(records)
        # 2 processed, then a truncation marker recorded in invalid.
        assert result.total == 2
        assert any("truncated" in " ".join(bad["errors"]).lower() for bad in result.invalid)

    def test_parsed_contact_to_row_shape(self):
        row = parsed_contact_to_row(_pc(external_id="e", name="N", phone="+14155550123", metadata={"k": "v"}))
        assert row == {
            "external_id": "e",
            "name": "N",
            "phone_number": "+14155550123",
            "contact_metadata": {"k": "v"},
        }

    def test_parse_uses_source(self):
        # End-to-end: a CSV source parsed + validated through the ONE loop.
        csv_text = "name,phone_number\nAlice,+14155550123\nBob,notaphone\n"
        parser = RecordParser(CompositeValidator([PhoneNumberValidator()]))
        result = parser.parse(CSVContactSource(), csv_text.encode("utf-8"))
        assert result.valid_count == 1 and result.invalid_count == 1
        assert result.valid[0].name == "Alice"


class TestSchemaSampleValue:
    """A schema's sample file shows a date/time example in the field's configured format."""

    @staticmethod
    def _cs():
        from core.services.contacts.contact_schema_service import ContactSchemaService

        return ContactSchemaService

    def test_configured_format_is_applied(self):
        f = SimpleNamespace(format="datetime", field_metadata={"datetime_format": "%m/%d/%Y %H:%M"})
        assert self._cs()._sample_value_for_field(f) == "01/31/2026 14:30"

    def test_am_pm_format(self):
        f = SimpleNamespace(format="datetime", field_metadata={"datetime_format": "%m/%d/%Y %I:%M %p"})
        assert self._cs()._sample_value_for_field(f) == "01/31/2026 02:30 PM"

    def test_date_no_format_is_iso_date(self):
        f = SimpleNamespace(format="date", field_metadata={})
        assert self._cs()._sample_value_for_field(f) == "2026-01-31"

    def test_datetime_no_format_is_iso(self):
        f = SimpleNamespace(format="datetime", field_metadata=None)
        assert self._cs()._sample_value_for_field(f) == "2026-01-31T14:30:00"

    def test_non_datetime_and_none_are_blank(self):
        assert self._cs()._sample_value_for_field(SimpleNamespace(format=None, field_metadata={})) == ""
        assert self._cs()._sample_value_for_field(None) == ""


class TestBuildContactValidator:
    """The ONE shared validator builder — composes rules from the destination context so the
    outbound file upload and the contact-create API validate the same way."""

    def _pc(self, name=None, phone=None, meta=None):
        return ParsedContact(external_id="x", name=name, phone_number=phone, metadata=meta or {})

    def test_require_phone_composes_phone_rule(self):
        v = build_contact_validator(require_phone=True)
        assert [type(r) for r in v._rules] == [PhoneNumberValidator]
        # dialing → a valid E.164 phone is required.
        assert v.validate(self._pc(phone="+14155550123")) == []
        assert v.validate(self._pc(name="Jo")) != []          # no phone → rejected
        assert v.validate(self._pc(phone="415")) != []         # bad phone → rejected

    def test_no_phone_composes_identity_rule(self):
        v = build_contact_validator()  # require_phone=False, no schema
        assert [type(r) for r in v._rules] == [RequiredIdentityValidator]
        assert v.validate(self._pc(name="Jo")) == []           # name-only is a valid contact
        assert v.validate(self._pc(phone="+14155550123")) == []
        assert v.validate(self._pc()) != []                    # neither → rejected

    def test_no_schema_skips_metadata_rule(self):
        assert len(build_contact_validator(None)._rules) == 1
        assert len(build_contact_validator([])._rules) == 1     # empty is falsy → skipped

    def test_schema_adds_metadata_rule(self):
        from core.models.schema_field import SchemaField

        fields = [SchemaField(field_name="city", type="string", is_mandatory=False, validators={})]
        v = build_contact_validator(fields)
        assert [type(r) for r in v._rules] == [RequiredIdentityValidator, SchemaMetadataValidator]

    def test_require_phone_with_schema(self):
        from core.models.schema_field import SchemaField

        fields = [SchemaField(field_name="city", type="string", is_mandatory=False, validators={})]
        v = build_contact_validator(fields, require_phone=True)
        assert [type(r) for r in v._rules] == [PhoneNumberValidator, SchemaMetadataValidator]
