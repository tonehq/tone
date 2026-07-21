"""Excel (``.xlsx``) contact source (openpyxl).

Reads the first (active) sheet of an ``.xlsx`` workbook, treats the first **non-empty**
row as the header, coerces every cell to a string WITHOUT mangling (numeric ids keep their
exact digits, integral values lose the ``.0``, datetimes become ISO strings), then hands
the resulting ``header -> value`` rows to the shared
:func:`~core.services.contact_ingestion.row_mapping.map_rows_to_parsed_contacts` — so a
``.xlsx`` produces the same ``ParsedContact``s as the equivalent CSV.

Only ``.xlsx`` (Office Open XML) is supported. Legacy ``.xls`` (OLE2) is rejected upstream
by :func:`~core.services.contact_ingestion.select_source_for_upload`, so this class is
never handed OLE2 bytes; a corrupt/non-xlsx blob raises from ``load_workbook`` and is
surfaced by the sync as a clean ``failed`` status.
"""

from __future__ import annotations

import io
from datetime import date, datetime, time
from typing import Dict, Iterator, List, Optional

import openpyxl

from core.services.contact_ingestion.base import ContactSource, ParsedContact
from core.services.contact_ingestion.row_mapping import map_rows_to_parsed_contacts


def _cell_to_str(value: object) -> str:
    """Coerce an openpyxl cell value to a string without mangling it.

    - ``None`` / blank -> ``""``
    - ``bool`` -> ``"TRUE"``/``"FALSE"`` (Excel semantics; before the int branch)
    - ``datetime``/``date``/``time`` -> ISO-8601 (matches the CSV scheduled_at path)
    - integral numbers (openpyxl reads numeric cells as int/float) -> plain int string,
      no ``.0`` so a phone number stored as a number maps to correct E.164
    - other floats -> fixed-point string (never scientific notation)
    - anything else -> ``str().strip()``
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # Fixed-point, then trim trailing zeros — avoids ``str()``'s scientific notation
        # for very large/small magnitudes (e.g. 1e20).
        text = format(value, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text
    return str(value).strip()


class ExcelContactSource(ContactSource):
    """Parse an uploaded ``.xlsx`` (first/active sheet only) into ``ParsedContact`` rows."""

    def parse(self, raw: bytes) -> Iterator[ParsedContact]:
        # read_only=True streams rows (memory-safe for large files); data_only=True reads
        # the last-computed value of formula cells rather than the formula text.
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:  # empty / sheetless workbook -> nothing to import.
                return
            yield from map_rows_to_parsed_contacts(self._iter_rows(ws))
        finally:
            wb.close()

    @staticmethod
    def _iter_rows(ws) -> Iterator[Dict[str, str]]:
        """Yield one ``header -> value`` dict per data row of the active sheet.

        The first **non-empty** row is the header (leading fully-blank rows — titles,
        spacer rows — are skipped). Duplicate header names mirror ``csv.DictReader``: the
        later column wins. Unnamed header columns are dropped (they'd normalize to an empty
        key the mapper ignores). Short data rows are padded with ``""``.
        """
        header: Optional[List[str]] = None
        for raw_row in ws.iter_rows(values_only=True):
            cells = [_cell_to_str(c) for c in raw_row]
            if header is None:
                if not any(cell for cell in cells):
                    continue  # skip leading blank rows before the header
                header = cells
                continue
            row: Dict[str, str] = {}
            for idx, key in enumerate(header):
                if not key:
                    continue  # unnamed header column -> drop
                row[key] = cells[idx] if idx < len(cells) else ""
            yield row
