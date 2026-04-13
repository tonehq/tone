"""Generate MODEL_BENCHMARKS.xlsx from benchmark data."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
provider_font = Font(bold=True, size=11)
wrap_align = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def style_sheet(ws, headers, data, col_widths):
    """Apply headers, data, and styling to a worksheet."""
    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            # Bold provider column (first column)
            if col_idx == 1:
                cell.font = provider_font

    # Column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"


# ═══════════════════════════════════════════
# Sheet 1: LLM Quality
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "LLM Quality"

llm_quality_headers = ["Provider", "Model", "MMLU", "HumanEval / Coding", "Arena ELO", "Context Window", "Notes"]
llm_quality_data = [
    # OpenAI
    ["OpenAI", "gpt-4o", "88.7%", "90.2%", "~1350", "128K", "Flagship multimodal model"],
    ["OpenAI", "gpt-4o-mini", "82.0%", "87.2%", "~1270", "128K", "Cost-efficient; outperforms Gemini Flash & Claude Haiku on MMLU"],
    ["OpenAI", "gpt-4-turbo", "86.5%", "67.0%", "~1310", "128K", "Predecessor to 4o; weaker on coding"],
    ["OpenAI", "gpt-4.1", "90.2%", "94.5%", "N/A", "1M", "SWE-bench 54.6%"],
    ["OpenAI", "gpt-5", "~90.2%", "~90.5%", "~1484", "128K–1M", "SWE-bench 74.9%; AIME 94.6%"],
    ["OpenAI", "o1", "91.8%", "89.3%", "~1370", "200K", "Reasoning model; MATH 96.4%"],
    ["OpenAI", "o3", "92.9%", "87.4%", "N/A", "200K", "Reasoning model; MATH 97.8%; GPQA 82.8%"],
    ["OpenAI", "o3-mini", "85.9%", "96.3%", "~1340", "200K", "Efficient reasoning; high variant: 97.6% HumanEval"],
    ["OpenAI", "o4-mini", "90.0%", "97.3%", "N/A", "200K", "High variant: 99.3% HumanEval, MATH 98.2%"],
    # Anthropic
    ["Anthropic", "claude-opus-4-6", "90.5%", "~97.8%", "1504 / 1549 (coding)", "1M", "#1 Arena overall & coding; GPQA 91.3%; SWE-bench 80.8%"],
    ["Anthropic", "claude-sonnet-4-6", "89.3%", "~98%", "1523 (coding)", "1M", "SWE-bench 79.6%; GPQA 74.1%"],
    ["Anthropic", "claude-opus-4-5", "~90.8%", "~93%", "~1465 (coding)", "200K", "SWE-bench 80.9%; MMLU-Pro 90%"],
    ["Anthropic", "claude-sonnet-4-5", "89.3%", "97.6%", "~1491 (coding)", "200K", "HumanEval leaderboard leader"],
    ["Anthropic", "claude-haiku-4-5", "90.8%", "N/A", "N/A", "200K", "3x faster than Sonnet; AIME 80.7%"],
    # Google
    ["Google", "gemini-2.5-pro", "88.6%", "N/A", "~1450", "1M", "AIME 2024: 92%; AIME 2025: 83%"],
    ["Google", "gemini-2.5-flash", "88.4%", "N/A", "~1380", "1M", "GPQA 82.8%; AIME 88%; cost-efficient"],
    ["Google", "gemini-2.5-flash-lite", "N/A", "N/A", "N/A", "1M", "~75% of Flash capability at 30% cost"],
    # Meta
    ["Meta", "llama-3.1-8b", "73.0%", "72.6%", "N/A", "128K", "Open-weight; smallest Llama 3.1"],
    ["Meta", "llama-3.3-70b", "86.0%", "88.0%", "~1250", "128K", "Improved coding over 3.1-70B"],
    # DeepSeek
    ["DeepSeek", "deepseek-chat (V3)", "87.1%", "82.6%", "~1320", "128K", "MoE architecture; GSM8K 89.3%"],
    ["DeepSeek", "deepseek-reasoner (R1)", "90.8%", "N/A", "~1360", "128K", "Chain-of-thought; AIME 2025: 87.5%; GPQA 81.0%"],
    # Others
    ["Alibaba", "qwen3-max (235B MoE)", "~83.9%", "N/A", "N/A", "262K (ext. to 1M)", "Open-weight; LiveCodeBench v6: 83.6%"],
    ["xAI", "grok-4", "86.6%", "~98%", "1491", "256K (Fast: 2M)", "GPQA Diamond 88% (all-time high); AIME ~93%"],
    ["Mistral", "mistral-large-3", "85.5%", "~92%", "~1290", "256K", "675B MoE; open-weight; strong multilingual"],
    ["Mistral", "mistral-small-3.1", "79.0%", "74.0%", "N/A", "128K", "24B params; fits 16GB RAM at Q4"],
    ["Perplexity", "sonar", "N/A", "N/A", "N/A", "127K", "Search-augmented; SimpleQA F-score 0.773"],
    ["Perplexity", "sonar-pro", "N/A", "N/A", "N/A", "200K", "Search-augmented; SimpleQA F-score 0.858"],
]

style_sheet(ws1, llm_quality_headers, llm_quality_data, [14, 28, 10, 18, 22, 16, 50])


# ═══════════════════════════════════════════
# Sheet 2: LLM Latency
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("LLM Latency")

llm_latency_headers = ["Provider", "Model", "TTFT (ms)", "Output Speed (tok/s)", "Total Latency (~500 tok)", "Notes"]
llm_latency_data = [
    # OpenAI
    ["OpenAI", "gpt-4o", "970", "102.1", "~6.5s", "Azure variant: 2.29s TTFT, 71.3 t/s"],
    ["OpenAI", "gpt-4o-mini", "3,360", "31.4", "~19s", "Slow on OpenAI API; Azure much faster: 1.67s TTFT, 99.4 t/s"],
    ["OpenAI", "gpt-4.1", "910", "117.2", "~5.2s", "Strong balanced perf"],
    ["OpenAI", "gpt-5 (high)", "87,660", "81.9", "~94s", "Reasoning model; very high TTFT due to extended thinking"],
    ["OpenAI", "o1", "33,970", "104.3", "~39s", "Reasoning model"],
    ["OpenAI", "o3-mini", "7,120", "137.6", "~10.7s", "Fast output for a reasoning model"],
    ["OpenAI", "o4-mini (high)", "43,880", "121.5", "~48s", "Reasoning model; high TTFT but strong throughput"],
    # Anthropic
    ["Anthropic", "claude-opus-4-6", "1,820", "38.6", "~15s", "Fast Mode available at 2.5x speed"],
    ["Anthropic", "claude-sonnet-4-6", "1,080–1,680", "41.6–44.1", "~13s", "Google Vertex lowest TTFT (1.08s)"],
    ["Anthropic", "claude-opus-4-5", "1,460", "52.2", "~11s", "Reasoning variant: 13.03s TTFT, 63.9 t/s"],
    ["Anthropic", "claude-sonnet-4-5", "1,350", "36.9", "~15s", "Below-avg output speed"],
    ["Anthropic", "claude-haiku-4-5", "639", "95.4", "~6s", "Fastest TTFT of any major model; up to 135 t/s"],
    # Google
    ["Google", "gemini-2.5-pro", "34,510", "112.0", "~39s", "Reasoning model; very high TTFT"],
    ["Google", "gemini-2.5-flash", "520", "220.0", "~2.8s", "Excellent speed/latency combo"],
    ["Google", "gemini-2.5-flash-lite", "290–410", "392.8", "~1.6s", "Ultra-fast; up to 887 t/s reported"],
    # DeepSeek
    ["DeepSeek", "deepseek-chat (V3)", "7,000", "34.0", "~21s", "Via DeepInfra: 820ms TTFT, 97 t/s"],
    ["DeepSeek", "deepseek-reasoner (R1)", "7,000+", "~34", "~22s+", "Reasoning model; similar throughput to V3"],
    # xAI
    ["xAI", "grok-4", "10,540", "50.9", "~20s", "Reasoning model"],
    ["xAI", "grok-4 (fast, non-reasoning)", "540", "124.3", "~4.6s", "Much faster non-reasoning variant"],
    # Mistral
    ["Mistral", "mistral-large-3", "4,410", "37.4", "~18s", "Notably slow output for its tier"],
    ["Mistral", "mistral-small-3.1", "280–910", "121.0", "~4.4s", "Google Vertex TTFT as low as 180ms"],
    # Perplexity
    ["Perplexity", "sonar-pro", "1,510", "123.7", "~5.5s", "Includes search/retrieval overhead"],
    ["Perplexity", "sonar-3.1-small", "160", "148.0", "N/A", "Very fast TTFT"],
    ["Perplexity", "sonar-3.1-large", "360", "58.0", "N/A", "Larger variant, slower throughput"],
    # Llama by provider
    ["Cerebras", "llama-3.1-8b", "N/A", "2,197.7", "<1s", "Wafer-scale engine; fastest by far"],
    ["Groq", "llama-3.1-8b", "N/A", "668.6", "~1s", "LPU architecture"],
    ["SambaNova", "llama-3.1-8b", "N/A", "634.5", "~1s", "Custom chip inference"],
    ["Fireworks", "llama-3.1-8b", "N/A", "~200", "~3s", "GPU-based, competitive TTFT (0.60s)"],
    ["Together", "llama-3.1-8b", "N/A", "~150", "~4s", "GPU-based provider"],
    ["Groq", "llama-3.3-70b", "800", "315.6", "~2.4s", "Speculative decoding variant: 1,665 t/s"],
    ["SambaNova", "llama-3.3-70b", "N/A", "294.1", "~2.5s", "Competitive with Groq"],
    ["Fireworks", "llama-3.3-70b", "600", "N/A", "N/A", "Tied for lowest TTFT"],
    ["Google Vertex", "llama-3.3-70b", "600", "N/A", "N/A", "Tied for lowest TTFT"],
]

style_sheet(ws2, llm_latency_headers, llm_latency_data, [14, 30, 14, 20, 22, 55])


# ═══════════════════════════════════════════
# Sheet 3: STT Benchmarks
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("STT Benchmarks")

stt_headers = ["Provider", "Model", "WER (AA-WER / LibriSpeech)", "Languages", "Streaming", "Latency", "Notes"]
stt_data = [
    ["Deepgram", "Nova-3", "5.4% AA-WER; ~2.2% LS clean", "47+", "Yes", "Sub-300ms", "Fastest commercial API (~113x RT). 54% WER reduction over nearest competitor"],
    ["Deepgram", "Nova-2-General", "5.5% AA-WER", "36", "Yes", "Sub-300ms", "Predecessor to Nova-3"],
    ["OpenAI", "whisper-1", "~4.2% AA-WER; ~2.7% LS clean", "57+", "No (batch)", "High", "Hosted Whisper Large v2; open-source base"],
    ["OpenAI", "gpt-4o-transcribe", "~4.1% AA-WER; ~2.5% LS clean", "50+", "Yes (Realtime API)", "Moderate", "Best accuracy among OpenAI STT models"],
    ["OpenAI", "gpt-4o-mini-transcribe", "N/A", "50+", "Yes (Realtime API)", "Lower", "Cost-optimized; slightly lower accuracy"],
    ["Groq", "whisper-large-v3-turbo", "~4.8% AA-WER", "57+", "No (batch)", "Ultra-low (216–252x RT)", "Fastest Whisper inference available"],
    ["Groq", "whisper-large-v3", "~2.7% LS clean", "57+", "No (batch)", "Very low (164x RT)", "Full Whisper v3 on LPU hardware"],
    ["AssemblyAI", "Universal-3-Pro", "3.2% AA-WER; 5.6% mean (26 datasets)", "99", "Yes (8.14% WER streaming)", "~61x RT", "Lowest streaming WER among major providers"],
    ["AssemblyAI", "Universal-2", "~2.4% LS clean", "99", "Yes", "N/A", "Broad language coverage"],
    ["Cartesia", "ink-whisper", "~19% WER (phone calls)", "57+ (trained on 98)", "Yes", "Fastest TTCT streaming", "Optimized for conversational/real-time; dynamic chunking"],
    ["Soniox", "stt-rt-v4", "~6.5% EN; 1.29% voice-agent benchmark", "60+", "Yes", "249ms median", "Purpose-built for real-time voice agents"],
    ["ElevenLabs", "scribe_v2", "2.3% AA-WER; 93.5% FLEURS accuracy", "90+", "Yes (Realtime variant)", "~150ms", "Lowest WER on Artificial Analysis leaderboard"],
    ["Gladia", "solaria-1", "~6% WER avg; 29% lower on conversational", "100+", "Yes", "103ms partial, 270ms avg", "First universal multilingual model; strong code-switching"],
    ["Google", "chirp_3", "~2.5% LS clean", "100+", "Yes", "N/A", "Speaker diarization and auto language detection"],
    ["Google", "chirp_2", "~9.8–11.6% AA-WER", "100+", "No (batch)", "N/A", "Batch-only; accuracy varies by dataset"],
    ["Azure", "Universal Language Model", "~13–23% WER (varies)", "100+", "Yes", "N/A", "Broad language support; MAI-Transcribe-1 (3.0% AA-WER) is newer"],
    ["Azure", "Whisper", "~4.2% AA-WER", "57+", "No (batch)", "N/A", "Azure-hosted OpenAI Whisper; custom fine-tuning available"],
    ["Speechmatics", "Enhanced Operating Point", "~4.3% AA-WER", "55+", "Yes", "+734ms over Deepgram baseline", "Higher accuracy than Standard at cost of speed"],
    ["NVIDIA", "Parakeet-TDT-0.6B-v2", "1.69% / 3.19% LS clean/other", "English only", "No (self-hosted)", "RTFx 3,386x (batch)", "Open-source; top of HuggingFace Open ASR Leaderboard"],
    ["NVIDIA", "Parakeet-TDT-0.6B-v3", "1.93% / 3.59% LS clean/other", "25 (EU languages)", "No (self-hosted)", "Similar to v2", "Multilingual extension; open-source"],
    ["Sarvam", "Saarika v2.5", "~13.58% WER (VISTAAR avg)", "11 Indian languages + EN", "Yes", "N/A", "Specialized for Indian languages; being deprecated"],
    ["SambaNova", "Whisper-Large-v3", "~2% clean / ~12% noisy", "57+", "No (batch)", "245x RT", "Fastest Whisper hosting alongside Groq"],
]

style_sheet(ws3, stt_headers, stt_data, [14, 24, 32, 20, 22, 24, 55])


# ═══════════════════════════════════════════
# Sheet 4: TTS Benchmarks
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("TTS Benchmarks")

tts_headers = ["Provider", "Model", "MOS / Quality", "Latency (TTFB)", "Languages", "Voices", "Streaming", "Notes"]
tts_data = [
    ["Cartesia", "Sonic 3", "ELO 1,054", "~40ms TTFA", "42", "Prebuilt + 3s voice cloning", "Yes", "60+ emotional tones; state-space architecture"],
    ["Cartesia", "Sonic 2", "N/A", "~90ms", "15", "Prebuilt + voice cloning", "Yes", "Predecessor to Sonic 3"],
    ["Cartesia", "Sonic Turbo", "N/A", "~40ms", "15", "Same as Sonic 2", "Yes", "Half-latency variant of Sonic 2"],
    ["ElevenLabs", "eleven_v3", "ELO ~1,108; MOS ~3.83", "Higher; not real-time suited", "32+", "380+ library voices", "Yes", "Most expressive model"],
    ["ElevenLabs", "eleven_multilingual_v2", "ELO 1,108", "~300ms+", "29", "380+", "Yes", "Flagship multilingual model"],
    ["ElevenLabs", "eleven_flash_v2_5", "N/A", "~75ms", "32", "380+", "Yes", "Lowest-latency ElevenLabs model"],
    ["ElevenLabs", "eleven_turbo_v2_5", "N/A", "~75–100ms", "32", "380+", "Yes", "Functionally equivalent to flash, slightly higher latency"],
    ["PlayHT", "PlayDialog", "N/A", "N/A", "30+", "Multiple + voice cloning", "Yes", "Flagship conversational/multi-turn model"],
    ["PlayHT", "Play3.0-mini", "N/A", "~143ms", "30+", "Multiple + voice cloning", "Yes", "28% faster inference than Play 2.0"],
    ["Deepgram", "Aura-2", "N/A", "~90ms; P95 <200ms", "7", "40+ English voices", "Yes", "Built in Rust; #1 in Coval real-time benchmarks"],
    ["Deepgram", "Aura", "N/A", "~250ms", "2", "20+", "Yes", "First-gen; English-centric"],
    ["OpenAI", "tts-1", "ELO 1,106", "~250ms", "57", "9 voices", "Yes", "Low-latency variant; MP3/Opus/AAC/FLAC/WAV/PCM"],
    ["OpenAI", "tts-1-hd", "N/A", "Higher than tts-1", "57", "9", "Yes", "Higher fidelity audio output"],
    ["OpenAI", "gpt-4o-mini-tts", "MOS >4.0", "~250ms", "32", "9 + steerable instructions", "Yes", "Instruction-steerable tone/style"],
    ["Hume AI", "Octave 1", "N/A", "~300ms+", "2", "Voice design from text prompts", "Yes", "Emotion-aware; voice cloning from 15s audio"],
    ["Hume AI", "Octave 2", "ELO 1,046", "<200ms", "11", "60+ prebuilt + prompt-designed", "Yes", "40% faster, 50% cheaper than Octave 1"],
    ["LMNT", "Blizzard", "N/A", "150–200ms", "24", "Multiple + 5s voice cloning", "Yes", "Mid-sentence voice switching; unlimited concurrency"],
    ["Rime", "Arcana", "N/A", "~80ms model / ~200ms cloud", "4", "40+ (18 EN accents)", "Yes", "Trained on real customer service data"],
    ["Rime", "Mist v2", "N/A", "~100ms on-prem / ~225ms cloud", "4", "Multiple", "Yes", "Powers tens of millions of production calls/month"],
    ["MiniMax", "Speech 2.8 HD", "ELO #1 (mid-2025)", "~300ms+", "Multiple", "Multiple + voice cloning", "Yes", "Studio-grade fidelity; top arena ELO for quality"],
    ["MiniMax", "Speech 2.8 Turbo", "N/A", "~200ms; <250ms", "Multiple", "Multiple + voice cloning", "Yes", "Speed-optimized variant"],
    ["Fish Audio", "Speech 1.5/1.6", "ELO 1,339", "<150ms", "13", "2M+ community; 10s cloning", "Yes", "Open-source weights available"],
    ["Resemble AI", "tts-v4 / Chatterbox", "63.75% preference vs ElevenLabs", "<200ms", "Multiple", "Custom voice cloning", "Yes", "350M-param turbo; open-source Chatterbox"],
    ["Neuphonic", "neu_hq", "N/A", "<25ms (cloud)", "4", "Multiple + 3s voice cloning", "Yes", "Ultra-low latency cloud TTS"],
    ["Inworld", "TTS-1.5 Max", "ELO 1,236 (#1 AA, Mar 2026)", "<250ms P90", "15", "Custom + voice cloning", "Yes", "30% more expressive, 40% lower WER vs prior gen"],
    ["Inworld", "TTS-1.5 Mini", "N/A", "<130ms P90", "15", "Custom", "Yes", "Lightweight low-latency variant"],
    ["Camb AI", "MARS8-Flash", "CER 5.67%; PQ 7.45", "~100ms", "150+", "Multiple + voice cloning", "Yes", "Ultra-low latency for real-time agents"],
    ["Camb AI", "MARS8-Pro", "0.87 speaker similarity; PQ 7.45", "800ms–2s", "150+", "Multiple + voice cloning", "Yes", "High-fidelity; suited for dubbing/production"],
    ["Google Cloud", "Chirp 3 HD", "N/A", "Up to 3.5s (long text)", "31", "8 styles per language", "Yes", "Advanced audio controls"],
    ["Google Cloud", "Studio", "N/A", "~200–400ms", "Limited (EN)", "~4", "Yes", "Designed for news/broadcast"],
    ["Google Cloud", "Neural2", "N/A", "~100–200ms", "40+", "100+", "Yes", "Good price/performance"],
    ["Azure", "Dragon HD", "N/A", "~354ms+", "19 GA languages", "30 HD voices", "Yes", "LM-enhanced context; auto-emotion detection"],
    ["Azure", "Neural", "N/A", "~100–200ms", "150+", "600+", "Yes", "Fastest Azure tier; massive coverage"],
    ["AWS Polly", "Generative", "ELO 1,060", "~100ms–1s", "20+ locales", "43 voices", "Yes", "Bidirectional streaming (Mar 2026)"],
    ["AWS Polly", "Neural", "N/A", "~100–300ms", "30+", "60+", "Yes", "Mid-tier quality"],
    ["AWS Polly", "Standard", "N/A", "~50–100ms", "40+", "60+", "Yes", "Lowest latency; concatenative synthesis"],
    ["Speechmatics", "TTS", "N/A", "~150ms", "EN (US, UK)", "Limited", "Yes", "Focus on reliability/clarity"],
    ["NVIDIA", "Magpie TTS", "N/A", "Self-hosted (varies)", "9", "2+ per language", "Yes", "Up to 64 concurrent streams"],
    ["NVIDIA", "FastPitch+HiFiGAN", "N/A", "Self-hosted (varies)", "1 (EN-US)", "Limited", "Yes", "Legacy mel-spectrogram + vocoder pipeline"],
    ["Groq", "Orpheus v1", "N/A", "N/A", "EN, Arabic", "Multiple", "Yes", "Community-driven; early stage"],
]

style_sheet(ws4, tts_headers, tts_data, [14, 24, 28, 26, 16, 28, 10, 50])


# ═══════════════════════════════════════════
# Save
# ═══════════════════════════════════════════
output_path = "/Users/thilak/Documents/Tone/docs/MODEL_BENCHMARKS.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
